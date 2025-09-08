# app.py
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, make_response
import sqlite3
import json
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import tempfile
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from pathlib import Path

from config import Config

app = Flask(__name__)
app.config.from_object(Config)

# 한글 폰트 등록
def register_korean_fonts():
    """한글 폰트를 ReportLab에 등록"""
    try:
        # 폰트 파일 경로
        font_dir = Path('fonts')
        
        # Noto Sans KR 폰트 파일 찾기
        noto_regular = None
        noto_bold = None
        
        # TTF 파일 찾기
        for font_file in font_dir.glob('*.ttf'):
            name = font_file.name.lower()
            if 'regular' in name or ('noto' in name and 'bold' not in name and 'light' not in name):
                noto_regular = font_file
            elif 'bold' in name:
                noto_bold = font_file
        
        # OTF 파일 찾기 (TTF가 없을 경우)
        if not noto_regular:
            for font_file in font_dir.glob('*.otf'):
                name = font_file.name.lower()
                if 'regular' in name or ('noto' in name and 'bold' not in name and 'light' not in name):
                    noto_regular = font_file
                elif 'bold' in name:
                    noto_bold = font_file
        
        # 폰트 등록
        if noto_regular and noto_regular.exists():
            pdfmetrics.registerFont(TTFont('NotoSansKR', str(noto_regular)))
            print(f"한글 폰트 등록 성공: {noto_regular}")
            
            if noto_bold and noto_bold.exists():
                pdfmetrics.registerFont(TTFont('NotoSansKR-Bold', str(noto_bold)))
                print(f"한글 굵은체 폰트 등록 성공: {noto_bold}")
                return 'NotoSansKR'
            else:
                return 'NotoSansKR'
        else:
            # DejaVu 폰트 사용 (Alpine Linux 기본 제공)
            try:
                # Alpine Linux DejaVu 폰트 경로
                pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/dejavu/DejaVuSans.ttf'))
                pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf'))
                print("DejaVu 폰트 등록 성공 (한글 일부 지원)")
                return 'DejaVuSans'
            except Exception as e:
                print("한글 폰트 파일을 찾을 수 없습니다. Helvetica 폰트를 사용합니다.")
                return 'Helvetica'
            
    except Exception as e:
        print(f"폰트 등록 중 오류: {e}")
        return 'Helvetica'

# 한글 폰트 등록 실행
KOREAN_FONT = register_korean_fonts()

# 데이터베이스 초기화
def init_db():
    try:
        conn = sqlite3.connect(app.config['DATABASE_PATH'])
        c = conn.cursor()
        
        print("데이터베이스 테이블 생성 중...")
        
        # 평가 영역 테이블
        c.execute('''CREATE TABLE IF NOT EXISTS categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            weight REAL NOT NULL,
            description TEXT,
            order_num INTEGER
        )''')
        
        # 평가 문항 테이블
        c.execute('''CREATE TABLE IF NOT EXISTS questions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category_id INTEGER,
            code TEXT NOT NULL,
            title TEXT NOT NULL,
            description TEXT,
            max_score INTEGER DEFAULT 5,
            order_num INTEGER,
            FOREIGN KEY (category_id) REFERENCES categories (id)
        )''')
        
        # 문항 선택지 테이블
        c.execute('''CREATE TABLE IF NOT EXISTS question_options (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            question_id INTEGER,
            score INTEGER,
            description TEXT,
            FOREIGN KEY (question_id) REFERENCES questions (id)
        )''')
        
        # 회사 테이블
        c.execute('''CREATE TABLE IF NOT EXISTS companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            industry TEXT,
            size TEXT,
            contact_person TEXT,
            contact_email TEXT,
            created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )''')
        
        # 평가 이력 테이블
        c.execute('''CREATE TABLE IF NOT EXISTS assessments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER,
            assessor_name TEXT,
            assessment_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            total_score REAL,
            maturity_level INTEGER,
            notes TEXT,
            status TEXT DEFAULT 'draft',
            last_modified TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completion_percentage INTEGER DEFAULT 0,
            FOREIGN KEY (company_id) REFERENCES companies (id)
        )''')
        
        # 평가 상세 결과 테이블
        c.execute('''CREATE TABLE IF NOT EXISTS assessment_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            assessment_id INTEGER,
            question_id INTEGER,
            score INTEGER,
            comment TEXT,
            FOREIGN KEY (assessment_id) REFERENCES assessments (id),
            FOREIGN KEY (question_id) REFERENCES questions (id)
        )''')
        
        # 평가 이력 추적 테이블
        c.execute('''CREATE TABLE IF NOT EXISTS assessment_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            assessment_id INTEGER,
            action_type TEXT,
            action_timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            user_info TEXT,
            questions_answered INTEGER,
            total_questions INTEGER,
            notes TEXT,
            FOREIGN KEY (assessment_id) REFERENCES assessments (id)
        )''')
        
        # 임시 저장 데이터 테이블
        c.execute('''CREATE TABLE IF NOT EXISTS assessment_drafts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            assessment_id INTEGER,
            question_id INTEGER,
            score INTEGER,
            comment TEXT,
            saved_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (assessment_id) REFERENCES assessments (id),
            FOREIGN KEY (question_id) REFERENCES questions (id)
        )''')
        
        # 기존 테이블에 새 컬럼 추가
        c.execute("PRAGMA table_info(assessment_results)")
        columns = [column[1] for column in c.fetchall()]
        if 'comment' not in columns:
            c.execute("ALTER TABLE assessment_results ADD COLUMN comment TEXT")
            print("assessment_results 테이블에 comment 컬럼 추가됨")
            
        # assessments 테이블에 새 컬럼 추가
        c.execute("PRAGMA table_info(assessments)")
        existing_columns = [column[1] for column in c.fetchall()]
        
        if 'status' not in existing_columns:
            c.execute("ALTER TABLE assessments ADD COLUMN status TEXT DEFAULT 'draft'")
            print("assessments 테이블에 status 컬럼 추가됨")
            
        if 'last_modified' not in existing_columns:
            c.execute("ALTER TABLE assessments ADD COLUMN last_modified TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
            print("assessments 테이블에 last_modified 컬럼 추가됨")
            
        if 'completion_percentage' not in existing_columns:
            c.execute("ALTER TABLE assessments ADD COLUMN completion_percentage INTEGER DEFAULT 0")
            print("assessments 테이블에 completion_percentage 컬럼 추가됨")
        
        conn.commit()
        print("데이터베이스 테이블 생성 완료")
        conn.close()
        
    except Exception as e:
        print(f"데이터베이스 초기화 오류: {e}")
        if conn:
            conn.close()

# 초기 데이터 삽입
def insert_initial_data():
    try:
        conn = sqlite3.connect(app.config['DATABASE_PATH'])
        c = conn.cursor()
        
        print("초기 데이터 확인 중...")
        
        # 카테고리 데이터 확인 및 삽입
        c.execute("SELECT COUNT(*) FROM categories")
        category_count = c.fetchone()[0]
        
        if category_count == 0:
            print("카테고리 초기 데이터 삽입 중...")
        else:
            print(f"카테고리 데이터 존재: {category_count}개")
        
        # 문항 데이터 확인 
        c.execute("SELECT COUNT(*) FROM questions")
        question_count = c.fetchone()[0]
        
        if question_count == 0:
            print("문항 초기 데이터 삽입 중...")
        else:
            print(f"문항 데이터 존재: {question_count}개 (기존 문항 보존)")
            # 기존 문항이 있으면 카테고리만 확인하고 문항은 건드리지 않음
            if category_count > 0:
                conn.close()
                return
        
        # 카테고리 삽입 (카테고리가 없을 경우에만)
        if category_count == 0:
            categories = [
                (1, '현행 프로세스 평가', 0.35, '생산계획수립, 스케줄생성, 작업지시, 실행, 분석 프로세스 평가', 1),
                (2, '데이터 준비도 평가', 0.35, '기준정보, 판매계획, 계획수립용데이터, 실행실적 데이터 평가', 2),
                (3, '관련 시스템 평가', 0.15, 'ERP, MES, 연동 시스템 등 IT 시스템 평가', 3),
                (4, '거버넌스 평가', 0.15, '인력, 조직, 의사결정체계, 경영진 지원 평가', 4)
            ]
            
            c.executemany("INSERT INTO categories (id, name, weight, description, order_num) VALUES (?, ?, ?, ?, ?)", categories)
        
        # 문항 삽입 (문항이 없을 경우에만)
        if question_count == 0:
            questions = [
            # 현행 프로세스 평가
            (1, 1, '1.1.1', '생산계획 수립 주기', '생산계획을 얼마나 자주 수립하는지 평가', 5, 1),
            (2, 1, '1.1.2', '계획 수립 시 고려하는 제약조건', '계획 수립 시 고려하는 제약조건의 범위', 5, 2),
            (3, 1, '1.1.3', '계획 수립 방법론', '계획 수립에 사용하는 방법론과 도구', 5, 3),
            (4, 1, '1.2.1', '상세 스케줄링 수준', '스케줄링의 상세화 정도', 5, 4),
            (5, 1, '1.2.2', '스케줄 최적화 고려사항', '스케줄 최적화 시 고려하는 요소들', 5, 5),
            (6, 1, '1.3.1', '작업지시 전달 방식', '작업지시를 전달하는 방식', 5, 6),
            (7, 1, '1.3.2', '실행 모니터링 및 추적', '생산 실행 과정의 모니터링 수준', 5, 7),
            
            # 데이터 준비도 평가
            (8, 2, '2.1.1', 'BOM 정확도', 'Bill of Materials의 정확도 수준', 5, 1),
            (9, 2, '2.1.2', '공정정보 완성도', 'Routing 정보의 완성도', 5, 2),
            (10, 2, '2.1.3', '설비/자원 정보 관리', '설비 및 자원 정보 관리 수준', 5, 3),
            (11, 2, '2.2.1', '수요예측/판매계획 정확도', '수요 예측의 정확도', 5, 4),
            (12, 2, '2.2.2', '재고정보 실시간성', '재고 정보의 실시간성', 5, 5),
            (13, 2, '2.3.1', '생산실적 수집 방식', '생산 실적 데이터 수집 방법', 5, 6),
            (14, 2, '2.3.2', '품질/불량 데이터 관리', '품질 및 불량 데이터 관리 수준', 5, 7),
            
            # 관련 시스템 평가
            (15, 3, '3.1.1', 'ERP 시스템 활용도', 'ERP 시스템의 활용 수준', 5, 1),
            (16, 3, '3.1.2', 'MES 구축 수준', 'MES 시스템 구축 및 활용 수준', 5, 2),
            (17, 3, '3.1.3', '시스템 간 연동 수준', '시스템 간 데이터 연동 수준', 5, 3),
            (18, 3, '3.2.1', 'IT 인프라 준비도', 'IT 인프라의 준비 상태', 5, 4),
            (19, 3, '3.2.2', '데이터 관리 체계', '데이터 관리 시스템 및 체계', 5, 5),
            (20, 3, '3.3.1', '시스템 확장성', '시스템의 확장 가능성', 5, 6),
            (21, 3, '3.3.2', '외부 시스템 연동 능력', '외부 시스템과의 연동 능력', 5, 7),
            
            # 거버넌스 평가
            (22, 4, '4.1.1', 'APS 관련 전담 인력', 'APS 관련 전담 인력 현황', 5, 1),
            (23, 4, '4.1.2', '관련 업무 담당자의 전문성', '담당자의 전문성 수준', 5, 2),
            (24, 4, '4.1.3', '교육 및 역량 개발 체계', '교육 및 역량 개발 프로그램', 5, 3),
            (25, 4, '4.2.1', 'APS 도입/운영 의사결정 체계', '의사결정 체계의 완성도', 5, 4),
            (26, 4, '4.2.2', '변화 관리 프로세스', '변화 관리 프로세스의 체계화', 5, 5),
            (27, 4, '4.3.1', '경영진의 APS 도입 의지', '경영진의 지원 의지', 5, 6),
            (28, 4, '4.3.2', '투자 계획 및 예산 확보', '투자 계획 및 예산 확보 상태', 5, 7)
        ]
        
            c.executemany("INSERT INTO questions (id, category_id, code, title, description, max_score, order_num) VALUES (?, ?, ?, ?, ?, ?, ?)", questions)
        
            # 문항별 선택지 삽입
            options_data = {
                1: {  # 생산계획 수립 주기
                    1: "불규칙적이며 수동으로 필요 시마다 수립",
                    2: "월 단위로 정기적 수립",
                    3: "주 단위로 정기적 수립",
                    4: "일 단위로 정기적 수립",
                    5: "실시간 또는 시간 단위로 동적 수립"
                },
                2: {  # 계획 수립 시 고려하는 제약조건
                    1: "기본 생산 용량만 고려",
                    2: "설비 용량 제약 고려",
                    3: "설비 용량 + 인력 제약 고려",
                    4: "설비 + 인력 + 자재 제약 고려",
                    5: "모든 제약조건(설비, 인력, 자재, 품질, 납기 등) 종합 고려"
                },
                8: {  # BOM 정확도
                    1: "60% 미만 (부정확한 정보 다수)",
                    2: "60-70% (기본적 정보만 관리)",
                    3: "70-85% (대부분 정확하나 일부 오류)",
                    4: "85-95% (높은 정확도, 정기 검증)",
                    5: "95% 이상 (매우 높은 정확도, 실시간 업데이트)"
                }
            }
            
            options = []
            for q_id in range(1, 29):
                for score in range(1, 6):
                    if q_id in options_data:
                        description = options_data[q_id][score]
                    else:
                        description = f"Level {score} - {['기본', '관리', '정의', '최적화', '혁신'][score-1]} 수준"
                    
                    options.append((q_id, score, description))
            
            c.executemany("INSERT INTO question_options (question_id, score, description) VALUES (?, ?, ?)", options)
        
        conn.commit()
        conn.close()
        print("초기 데이터 삽입 완료")
        
    except Exception as e:
        print(f"초기 데이터 삽입 오류: {e}")
        if conn:
            conn.close()

# 성숙도 레벨 계산
def calculate_maturity_level(total_score):
    max_score = 140  # 28문항 × 5점
    percentage = (total_score / max_score) * 100
    
    if percentage < 40:
        return 1
    elif percentage < 60:
        return 2
    elif percentage < 80:
        return 3
    elif percentage < 91:
        return 4
    else:
        return 5

# 라우트 정의
@app.route('/health')
def health_check():
    """헬스체크 엔드포인트"""
    return {'status': 'healthy', 'message': 'APS Assessment System is running'}, 200

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/companies')
def companies():
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    c.execute('''SELECT c.*, 
                        COUNT(a.id) as assessment_count,
                        draft_assessments.draft_id,
                        draft_assessments.completion_percentage
                 FROM companies c 
                 LEFT JOIN assessments a ON c.id = a.company_id 
                 LEFT JOIN (
                     SELECT company_id, id as draft_id, completion_percentage 
                     FROM assessments 
                     WHERE status = 'draft'
                 ) draft_assessments ON c.id = draft_assessments.company_id
                 GROUP BY c.id
                 ORDER BY c.created_date DESC''')
    companies_data = c.fetchall()
    conn.close()
    return render_template('companies.html', companies=companies_data)

@app.route('/company/new', methods=['GET', 'POST'])
def new_company():
    if request.method == 'POST':
        conn = sqlite3.connect(app.config['DATABASE_PATH'])
        c = conn.cursor()
        c.execute('''INSERT INTO companies (name, industry, size, contact_person, contact_email)
                     VALUES (?, ?, ?, ?, ?)''',
                  (request.form['name'], request.form['industry'], request.form['size'],
                   request.form['contact_person'], request.form['contact_email']))
        conn.commit()
        conn.close()
        flash('회사가 성공적으로 등록되었습니다.')
        return redirect(url_for('companies'))
    return render_template('company_form.html')

@app.route('/assessment/new/<int:company_id>')
def new_assessment(company_id):
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    
    # URL에서 assessment_id 파라미터 확인 (계속하기용)
    assessment_id = request.args.get('assessment_id')
    
    if not assessment_id:
        # 해당 회사의 기존 임시저장 평가 확인
        c.execute("SELECT id FROM assessments WHERE company_id = ? AND status = 'draft'", (company_id,))
        existing_draft = c.fetchone()
        
        if existing_draft:
            # 기존 임시저장이 있으면 해당 평가로 리다이렉트
            conn.close()
            return redirect(url_for('continue_assessment', assessment_id=existing_draft[0]))
    
    # 회사 정보
    c.execute("SELECT * FROM companies WHERE id = ?", (company_id,))
    company = c.fetchone()
    
    # 카테고리별 질문
    c.execute('''SELECT c.*, q.* FROM categories c
                 LEFT JOIN questions q ON c.id = q.category_id
                 ORDER BY c.order_num, q.order_num''')
    data = c.fetchall()
    
    # 질문별 선택지
    c.execute('''SELECT qo.question_id, qo.score, qo.description 
                 FROM question_options qo
                 ORDER BY qo.question_id, qo.score''')
    options_data = c.fetchall()
    
    # 기존 임시저장 데이터 로드 (assessment_id가 있는 경우)
    existing_answers = {}
    existing_assessment = None
    if assessment_id:
        c.execute("SELECT * FROM assessments WHERE id = ? AND status = 'draft'", (assessment_id,))
        existing_assessment = c.fetchone()
        
        if existing_assessment:
            # 기존 답변 로드
            c.execute("SELECT question_id, score, comment FROM assessment_results WHERE assessment_id = ?", (assessment_id,))
            results = c.fetchall()
            for result in results:
                existing_answers[result[0]] = {
                    'score': result[1],
                    'opinion': result[2] or ''
                }
    
    conn.close()
    
    # 데이터 구조화
    categories = {}
    options = {}
    
    for row in data:
        cat_id = row[0]
        if cat_id not in categories:
            categories[cat_id] = {
                'id': row[0], 'name': row[1], 'weight': row[2], 
                'description': row[3], 'questions': []
            }
        if row[5]:  # question exists
            categories[cat_id]['questions'].append({
                'id': row[5], 'code': row[7], 'title': row[8], 'description': row[9]
            })
    
    for option in options_data:
        q_id = option[0]
        if q_id not in options:
            options[q_id] = []
        options[q_id].append({'score': option[1], 'description': option[2]})
    
    return render_template('assessment_form.html', company=company, 
                         categories=categories, options=options,
                         existing_answers=existing_answers, 
                         existing_assessment=existing_assessment)

# 임시저장 관련 라우트들
@app.route('/assessment/save_draft', methods=['POST'])
def save_draft():
    """평가 임시저장"""
    try:
        data = request.get_json()
        company_id = data.get('company_id')
        assessor_name = data.get('assessor_name')
        assessment_id = data.get('assessment_id')  # 기존 평가 ID (있는 경우)
        answers = data.get('answers', {})
        notes = data.get('notes', '')
        
        conn = sqlite3.connect(app.config['DATABASE_PATH'])
        c = conn.cursor()
        
        # 새 평가인지 기존 평가 수정인지 확인
        if assessment_id:
            # 기존 평가 업데이트
            c.execute("SELECT id FROM assessments WHERE id = ? AND status = 'draft'", (assessment_id,))
            if not c.fetchone():
                return {'status': 'error', 'message': '수정할 수 없는 평가입니다.'}, 400
        else:
            # 새 평가 생성
            c.execute('''INSERT INTO assessments (company_id, assessor_name, notes, status, 
                         last_modified, completion_percentage)
                         VALUES (?, ?, ?, 'draft', CURRENT_TIMESTAMP, ?)''',
                      (company_id, assessor_name, notes, 0))
            assessment_id = c.lastrowid
            
            # 평가 이력 추가
            c.execute('''INSERT INTO assessment_history (assessment_id, action_type, user_info,
                         questions_answered, total_questions, notes)
                         VALUES (?, 'created', ?, 0, 28, '평가 시작')''',
                      (assessment_id, assessor_name))
        
        # 기존 임시저장 데이터 삭제
        c.execute("DELETE FROM assessment_drafts WHERE assessment_id = ?", (assessment_id,))
        
        # 새 임시저장 데이터 삽입
        questions_answered = 0
        for question_id, answer_data in answers.items():
            if answer_data.get('score'):
                score = answer_data.get('score')
                comment = answer_data.get('comment', '')
                c.execute('''INSERT INTO assessment_drafts (assessment_id, question_id, score, comment)
                             VALUES (?, ?, ?, ?)''', (assessment_id, question_id, score, comment))
                questions_answered += 1
        
        # 진행률 계산 및 업데이트
        completion_percentage = int((questions_answered / 28) * 100)
        c.execute('''UPDATE assessments SET completion_percentage = ?, last_modified = CURRENT_TIMESTAMP,
                     notes = ?, assessor_name = ? WHERE id = ?''', (completion_percentage, notes, assessor_name or '', assessment_id))
        
        # 이력 추가
        c.execute('''INSERT INTO assessment_history (assessment_id, action_type, user_info,
                     questions_answered, total_questions, notes)
                     VALUES (?, 'saved_draft', ?, ?, 28, '임시저장')''',
                  (assessment_id, assessor_name, questions_answered))
        
        conn.commit()
        conn.close()
        
        return {
            'status': 'success', 
            'assessment_id': assessment_id,
            'completion_percentage': completion_percentage,
            'message': f'임시저장 완료 ({questions_answered}/28 문항)'
        }
        
    except Exception as e:
        return {'status': 'error', 'message': str(e)}, 500

@app.route('/assessment/load_draft/<int:assessment_id>')
def load_draft(assessment_id):
    """임시저장된 평가 불러오기"""
    try:
        conn = sqlite3.connect(app.config['DATABASE_PATH'])
        c = conn.cursor()
        
        # 평가 정보 확인
        c.execute('''SELECT * FROM assessments WHERE id = ? AND status = 'draft' ''', (assessment_id,))
        assessment = c.fetchone()
        
        if not assessment:
            return {'status': 'error', 'message': '임시저장된 평가를 찾을 수 없습니다.'}, 404
        
        # 임시저장 데이터 불러오기
        c.execute('''SELECT question_id, score, comment FROM assessment_drafts 
                     WHERE assessment_id = ?''', (assessment_id,))
        draft_data = c.fetchall()
        
        conn.close()
        
        # 데이터 형식 변환
        answers = {}
        for question_id, score, comment in draft_data:
            answers[str(question_id)] = {
                'score': score,
                'comment': comment or ''
            }
        
        return {
            'status': 'success',
            'assessment_id': assessment_id,
            'answers': answers,
            'notes': assessment[6] or '',
            'assessor_name': assessment[2] or '',
            'completion_percentage': assessment[9] or 0
        }
        
    except Exception as e:
        return {'status': 'error', 'message': str(e)}, 500

@app.route('/assessment/delete_draft/<int:assessment_id>', methods=['DELETE'])
def delete_draft(assessment_id):
    """임시저장된 평가 삭제"""
    try:
        conn = sqlite3.connect(app.config['DATABASE_PATH'])
        c = conn.cursor()
        
        # draft 상태인지 확인
        c.execute("SELECT status FROM assessments WHERE id = ?", (assessment_id,))
        result = c.fetchone()
        
        if not result:
            return {'status': 'error', 'message': '평가를 찾을 수 없습니다.'}, 404
            
        if result[0] != 'draft':
            return {'status': 'error', 'message': '완료된 평가는 삭제할 수 없습니다.'}, 400
        
        # 관련 데이터 모두 삭제
        c.execute("DELETE FROM assessment_drafts WHERE assessment_id = ?", (assessment_id,))
        c.execute("DELETE FROM assessment_history WHERE assessment_id = ?", (assessment_id,))
        c.execute("DELETE FROM assessments WHERE id = ?", (assessment_id,))
        
        conn.commit()
        conn.close()
        
        return {'status': 'success', 'message': '임시저장된 평가가 삭제되었습니다.'}
        
    except Exception as e:
        return {'status': 'error', 'message': str(e)}, 500

@app.route('/assessment/continue/<int:assessment_id>')
def continue_assessment(assessment_id):
    """임시저장된 평가 계속하기"""
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    
    # 평가 정보와 회사 정보 조회
    c.execute('''SELECT a.company_id, c.name, c.industry 
                 FROM assessments a
                 JOIN companies c ON a.company_id = c.id
                 WHERE a.id = ? AND a.status = 'draft' ''', (assessment_id,))
    
    result = c.fetchone()
    conn.close()
    
    if not result:
        flash('임시저장된 평가를 찾을 수 없습니다.')
        return redirect(url_for('assessments'))
    
    # new_assessment로 리다이렉트하면서 assessment_id 전달
    return redirect(url_for('new_assessment', company_id=result[0]) + f'?assessment_id={assessment_id}')

@app.route('/assessment/submit', methods=['POST'])
def submit_assessment():
    company_id = request.form['company_id']
    assessor_name = request.form['assessor_name']
    notes = request.form.get('notes', '')
    assessment_id = request.form.get('assessment_id')  # 기존 임시저장 ID가 있을 경우
    
    # 점수 계산 및 주관식 답변 수집
    total_score = 0
    results = []
    comments = {}
    
    # 주관식 답변 수집
    for key, value in request.form.items():
        if key.startswith('comment_'):
            question_id = int(key.split('_')[1])
            if value.strip():  # 빈 문자열이 아닌 경우만 저장
                comments[question_id] = value.strip()
    
    # 점수 수집
    for key, value in request.form.items():
        if key.startswith('question_'):
            question_id = int(key.split('_')[1])
            score = int(value)
            total_score += score
            comment = comments.get(question_id, '')
            results.append((question_id, score, comment))
    
    maturity_level = calculate_maturity_level(total_score)
    
    # 데이터베이스 저장
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    
    if assessment_id and assessment_id.isdigit():
        # 기존 임시저장을 완료로 업데이트
        assessment_id = int(assessment_id)
        c.execute('''UPDATE assessments SET total_score = ?, maturity_level = ?, notes = ?,
                     status = 'completed', last_modified = CURRENT_TIMESTAMP, completion_percentage = 100,
                     assessor_name = ? WHERE id = ?''',
                  (total_score, maturity_level, notes, assessor_name, assessment_id))
        
        # 기존 임시저장 데이터 삭제
        c.execute("DELETE FROM assessment_drafts WHERE assessment_id = ?", (assessment_id,))
        
        # 이력 추가
        c.execute('''INSERT INTO assessment_history (assessment_id, action_type, user_info,
                     questions_answered, total_questions, notes)
                     VALUES (?, 'completed', ?, 28, 28, '평가 완료')''',
                  (assessment_id, assessor_name))
    else:
        # 새 평가 생성 (완료 상태로)
        c.execute('''INSERT INTO assessments (company_id, assessor_name, total_score, maturity_level, notes,
                     status, completion_percentage, last_modified)
                     VALUES (?, ?, ?, ?, ?, 'completed', 100, CURRENT_TIMESTAMP)''',
                  (company_id, assessor_name, total_score, maturity_level, notes))
        assessment_id = c.lastrowid
        
        # 이력 추가
        c.execute('''INSERT INTO assessment_history (assessment_id, action_type, user_info,
                     questions_answered, total_questions, notes)
                     VALUES (?, 'completed', ?, 28, 28, '평가 완료')''',
                  (assessment_id, assessor_name))
    
    # 상세 결과 저장 (기존 데이터 삭제 후 재삽입)
    c.execute("DELETE FROM assessment_results WHERE assessment_id = ?", (assessment_id,))
    for question_id, score, comment in results:
        c.execute('''INSERT INTO assessment_results (assessment_id, question_id, score, comment)
                     VALUES (?, ?, ?, ?)''', (assessment_id, question_id, score, comment))
    
    conn.commit()
    conn.close()
    
    flash(f'평가가 완료되었습니다. 총점: {total_score}/140, 성숙도 Level: {maturity_level}')
    return redirect(url_for('assessment_detail', assessment_id=assessment_id))

@app.route('/assessment/<int:assessment_id>')
def assessment_detail(assessment_id):
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    
    # 평가 기본 정보
    c.execute('''SELECT a.*, c.name as company_name FROM assessments a
                 JOIN companies c ON a.company_id = c.id
                 WHERE a.id = ?''', (assessment_id,))
    assessment = c.fetchone()
    
    # 카테고리별 점수
    c.execute('''SELECT cat.id, cat.name, cat.weight, SUM(ar.score) as category_score,
                        COUNT(ar.score) * 5 as max_category_score
                 FROM categories cat
                 JOIN questions q ON cat.id = q.category_id
                 JOIN assessment_results ar ON q.id = ar.question_id
                 WHERE ar.assessment_id = ?
                 GROUP BY cat.id, cat.name, cat.weight
                 ORDER BY cat.order_num''', (assessment_id,))
    category_scores = c.fetchall()
    
    # 상세 결과
    c.execute('''SELECT q.code, q.title, ar.score, qo.description, ar.comment
                 FROM assessment_results ar
                 JOIN questions q ON ar.question_id = q.id
                 JOIN question_options qo ON q.id = qo.question_id AND ar.score = qo.score
                 WHERE ar.assessment_id = ?
                 ORDER BY q.category_id, q.order_num''', (assessment_id,))
    detailed_results = c.fetchall()
    
    conn.close()
    
    return render_template('assessment_detail.html', assessment=assessment,
                         category_scores=category_scores, detailed_results=detailed_results)

@app.route('/assessments')
def assessments():
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    c.execute('''SELECT a.id, c.name as company_name, a.assessor_name, 
                        a.assessment_date, a.total_score, a.maturity_level,
                        a.status, a.completion_percentage, a.last_modified
                 FROM assessments a
                 JOIN companies c ON a.company_id = c.id
                 ORDER BY a.status ASC, a.last_modified DESC''')
    assessments_data = c.fetchall()
    conn.close()
    return render_template('assessments.html', assessments=assessments_data)

@app.route('/assessment_history')
def assessment_history():
    """평가 이력 관리 페이지"""
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    
    # 전체 평가 통계
    c.execute('''SELECT 
                   COUNT(*) as total,
                   COUNT(CASE WHEN status = 'draft' THEN 1 END) as draft_count,
                   COUNT(CASE WHEN status = 'completed' THEN 1 END) as completed_count
                 FROM assessments''')
    stats = c.fetchone()
    
    # 평가자별 활동 현황
    c.execute('''SELECT 
                   assessor_name,
                   COUNT(*) as total_assessments,
                   COUNT(CASE WHEN status = 'completed' THEN 1 END) as completed,
                   COUNT(CASE WHEN status = 'draft' THEN 1 END) as draft,
                   MAX(last_modified) as last_activity
                 FROM assessments 
                 GROUP BY assessor_name
                 ORDER BY last_activity DESC''')
    assessor_stats = c.fetchall()
    
    # 최근 활동 이력
    c.execute('''SELECT 
                   h.action_timestamp,
                   h.action_type,
                   h.user_info,
                   c.name as company_name,
                   h.questions_answered,
                   h.total_questions,
                   h.notes,
                   a.id as assessment_id
                 FROM assessment_history h
                 JOIN assessments a ON h.assessment_id = a.id
                 JOIN companies c ON a.company_id = c.id
                 ORDER BY h.action_timestamp DESC
                 LIMIT 50''')
    recent_activities = c.fetchall()
    
    # 월별 완료 통계 (최근 6개월)
    c.execute('''SELECT 
                   strftime('%Y-%m', assessment_date) as month,
                   COUNT(*) as completed_count
                 FROM assessments 
                 WHERE status = 'completed' 
                   AND assessment_date >= date('now', '-6 months')
                 GROUP BY strftime('%Y-%m', assessment_date)
                 ORDER BY month DESC''')
    monthly_stats = c.fetchall()
    
    conn.close()
    
    return render_template('assessment_history.html', 
                         stats=stats,
                         assessor_stats=assessor_stats,
                         recent_activities=recent_activities,
                         monthly_stats=monthly_stats)

@app.route('/api/assessment/<int:assessment_id>/chart')
def assessment_chart_data(assessment_id):
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    
    c.execute('''SELECT cat.name, SUM(ar.score) as score, COUNT(ar.score) * 5 as max_score
                 FROM categories cat
                 JOIN questions q ON cat.id = q.category_id
                 JOIN assessment_results ar ON q.id = ar.question_id
                 WHERE ar.assessment_id = ?
                 GROUP BY cat.id, cat.name
                 ORDER BY cat.order_num''', (assessment_id,))
    
    data = c.fetchall()
    conn.close()
    
    chart_data = {
        'categories': [row[0] for row in data],
        'scores': [row[1] for row in data],
        'maxScores': [row[2] for row in data],
        'percentages': [round((row[1]/row[2])*100, 1) for row in data]
    }
    
    return jsonify(chart_data)

@app.route('/api/assessment/<int:assessment_id>/category/<int:category_id>/detail')
def assessment_category_detail(assessment_id, category_id):
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    
    c.execute('''SELECT q.title, ar.score, cat.name as category_name
                 FROM questions q
                 JOIN assessment_results ar ON q.id = ar.question_id
                 JOIN categories cat ON q.category_id = cat.id
                 WHERE ar.assessment_id = ? AND q.category_id = ?
                 ORDER BY q.order_num''', (assessment_id, category_id))
    
    data = c.fetchall()
    conn.close()
    
    if not data:
        return jsonify({'error': 'No data found'}), 404
    
    detail_data = {
        'categoryName': data[0][2],
        'questions': [row[0] for row in data],
        'scores': [row[1] for row in data],
        'maxScore': 5,
        'percentages': [row[1] * 20 for row in data]  # Convert 1-5 score to 0-100 percentage
    }
    
    return jsonify(detail_data)

@app.route('/questions')
def questions():
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    c.execute('''SELECT q.id, c.name as category_name, q.code, q.title, q.description
                 FROM questions q
                 JOIN categories c ON q.category_id = c.id
                 ORDER BY c.order_num, q.order_num''')
    questions_data = c.fetchall()
    conn.close()
    return render_template('questions.html', questions=questions_data)

@app.route('/question/<int:question_id>/edit', methods=['GET', 'POST'])
def edit_question(question_id):
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    
    if request.method == 'POST':
        # 문항 정보 업데이트
        c.execute('''UPDATE questions SET code = ?, title = ?, description = ? 
                     WHERE id = ?''',
                  (request.form['code'], request.form['title'], 
                   request.form['description'], question_id))
        
        # 선택지 업데이트
        for score in range(1, 6):
            option_desc = request.form.get(f'option_{score}', '')
            c.execute('''UPDATE question_options SET description = ? 
                         WHERE question_id = ? AND score = ?''',
                      (option_desc, question_id, score))
        
        conn.commit()
        conn.close()
        flash('문항이 성공적으로 수정되었습니다.')
        return redirect(url_for('questions'))
    
    # GET 요청 - 수정 폼 표시
    # 문항 정보 조회
    c.execute('''SELECT q.*, c.name as category_name, c.id as category_id
                 FROM questions q
                 JOIN categories c ON q.category_id = c.id
                 WHERE q.id = ?''', (question_id,))
    question = c.fetchone()
    
    # 카테고리 목록 조회
    c.execute('SELECT id, name FROM categories ORDER BY order_num')
    categories = c.fetchall()
    
    # 선택지 조회
    c.execute('''SELECT score, description FROM question_options 
                 WHERE question_id = ? ORDER BY score''', (question_id,))
    options = c.fetchall()
    
    conn.close()
    
    return render_template('question_edit.html', question=question, 
                         categories=categories, options=options)

@app.route('/question/<int:question_id>/delete', methods=['POST'])
def delete_question(question_id):
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    
    # 선택지 먼저 삭제
    c.execute('DELETE FROM question_options WHERE question_id = ?', (question_id,))
    
    # 평가 결과 삭제 (만약 있다면)
    c.execute('DELETE FROM assessment_results WHERE question_id = ?', (question_id,))
    
    # 문항 삭제
    c.execute('DELETE FROM questions WHERE id = ?', (question_id,))
    
    conn.commit()
    conn.close()
    
    flash('문항이 성공적으로 삭제되었습니다.')
    return redirect(url_for('questions'))

@app.route('/question/new', methods=['GET', 'POST'])
def new_question():
    if request.method == 'POST':
        conn = sqlite3.connect(app.config['DATABASE_PATH'])
        c = conn.cursor()
        
        # 새 문항 추가
        c.execute('''INSERT INTO questions (category_id, code, title, description, max_score, order_num)
                     VALUES (?, ?, ?, ?, 5, 
                     (SELECT COALESCE(MAX(order_num), 0) + 1 FROM questions WHERE category_id = ?))''',
                  (request.form['category_id'], request.form['code'], request.form['title'],
                   request.form['description'], request.form['category_id']))
        
        question_id = c.lastrowid
        
        # 선택지 추가
        for score in range(1, 6):
            option_desc = request.form.get(f'option_{score}', f'Level {score}')
            c.execute('''INSERT INTO question_options (question_id, score, description)
                         VALUES (?, ?, ?)''', (question_id, score, option_desc))
        
        conn.commit()
        conn.close()
        
        flash('새 문항이 성공적으로 추가되었습니다.')
        return redirect(url_for('questions'))
    
    # GET 요청 - 새 문항 폼 표시
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    c.execute('SELECT id, name FROM categories ORDER BY order_num')
    categories = c.fetchall()
    conn.close()
    
    return render_template('question_new.html', categories=categories)

@app.route('/categories')
def categories():
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    c.execute('''SELECT c.*, COUNT(q.id) as question_count
                 FROM categories c
                 LEFT JOIN questions q ON c.id = q.category_id
                 GROUP BY c.id
                 ORDER BY c.order_num''')
    categories_data = c.fetchall()
    conn.close()
    return render_template('categories.html', categories=categories_data)

@app.route('/category/<int:category_id>/edit', methods=['GET', 'POST'])
def edit_category(category_id):
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    
    if request.method == 'POST':
        c.execute('''UPDATE categories SET name = ?, weight = ?, description = ?
                     WHERE id = ?''',
                  (request.form['name'], float(request.form['weight']), 
                   request.form['description'], category_id))
        conn.commit()
        conn.close()
        flash('카테고리가 성공적으로 수정되었습니다.')
        return redirect(url_for('categories'))
    
    # GET 요청
    c.execute('SELECT * FROM categories WHERE id = ?', (category_id,))
    category = c.fetchone()
    conn.close()
    
    return render_template('category_edit.html', category=category)

@app.route('/category/new', methods=['GET', 'POST'])
def new_category():
    if request.method == 'POST':
        conn = sqlite3.connect(app.config['DATABASE_PATH'])
        c = conn.cursor()
        
        # 새 카테고리의 order_num 계산 (기존 최대값 + 1)
        c.execute('SELECT COALESCE(MAX(order_num), 0) + 1 FROM categories')
        new_order_num = c.fetchone()[0]
        
        c.execute('''INSERT INTO categories (name, weight, description, order_num)
                     VALUES (?, ?, ?, ?)''',
                  (request.form['name'], float(request.form['weight']), 
                   request.form['description'], new_order_num))
        conn.commit()
        conn.close()
        flash('새 카테고리가 성공적으로 추가되었습니다.')
        return redirect(url_for('categories'))
    
    return render_template('category_new.html')

@app.route('/category/<int:category_id>/delete', methods=['POST'])
def delete_category(category_id):
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    
    # 카테고리에 연결된 문항이 있는지 확인
    c.execute('SELECT COUNT(*) FROM questions WHERE category_id = ?', (category_id,))
    question_count = c.fetchone()[0]
    
    if question_count > 0:
        flash(f'이 카테고리에는 {question_count}개의 문항이 연결되어 있어 삭제할 수 없습니다. 먼저 연결된 문항들을 삭제하거나 다른 카테고리로 이동해주세요.')
        conn.close()
        return redirect(url_for('categories'))
    
    # 카테고리 삭제
    c.execute('DELETE FROM categories WHERE id = ?', (category_id,))
    conn.commit()
    conn.close()
    
    flash('카테고리가 성공적으로 삭제되었습니다.')
    return redirect(url_for('categories'))

@app.route('/questions/export')
def export_questions():
    """평가 문항을 Excel 파일로 내보내기"""
    conn = sqlite3.connect(app.config['DATABASE_PATH'])
    c = conn.cursor()
    
    # 문항과 선택지 데이터 조회
    c.execute('''SELECT q.id, c.name as category_name, q.code, q.title, q.description,
                        qo.score, qo.description as option_desc
                 FROM questions q
                 JOIN categories c ON q.category_id = c.id
                 JOIN question_options qo ON q.id = qo.question_id
                 ORDER BY c.order_num, q.order_num, qo.score''')
    data = c.fetchall()
    conn.close()
    
    # Excel 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "평가문항"
    
    # 헤더 스타일 설정
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 작성
    headers = ["문항ID", "카테고리", "문항코드", "문항제목", "문항설명", "점수1", "점수2", "점수3", "점수4", "점수5"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 데이터 구조화 (문항별로 그룹화)
    questions_dict = {}
    for row in data:
        q_id = row[0]
        if q_id not in questions_dict:
            questions_dict[q_id] = {
                'id': row[0], 'category': row[1], 'code': row[2], 
                'title': row[3], 'description': row[4], 'options': {}
            }
        questions_dict[q_id]['options'][row[5]] = row[6]
    
    # 데이터 행 작성
    row_num = 2
    for q_data in questions_dict.values():
        ws.cell(row=row_num, column=1, value=q_data['id'])
        ws.cell(row=row_num, column=2, value=q_data['category'])
        ws.cell(row=row_num, column=3, value=q_data['code'])
        ws.cell(row=row_num, column=4, value=q_data['title'])
        ws.cell(row=row_num, column=5, value=q_data['description'])
        
        # 선택지 (점수 1-5)
        for score in range(1, 6):
            ws.cell(row=row_num, column=5 + score, value=q_data['options'].get(score, ''))
        
        row_num += 1
    
    # 열 너비 자동 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # 메모리에 파일 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 파일명 생성
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"평가문항_{timestamp}.xlsx"
    
    return send_file(output, 
                     as_attachment=True, 
                     download_name=filename, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/questions/import', methods=['POST'])
def import_questions():
    """Excel 파일에서 평가 문항 가져오기"""
    if 'file' not in request.files:
        flash('파일이 선택되지 않았습니다.')
        return redirect(url_for('questions'))
    
    file = request.files['file']
    if file.filename == '':
        flash('파일이 선택되지 않았습니다.')
        return redirect(url_for('questions'))
    
    if not file.filename.endswith('.xlsx'):
        flash('Excel 파일(.xlsx)만 업로드 가능합니다.')
        return redirect(url_for('questions'))
    
    try:
        # 임시 파일로 저장
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.save(tmp.name)
            
            # Excel 파일 읽기
            wb = load_workbook(tmp.name)
            ws = wb.active
            
            conn = sqlite3.connect(app.config['DATABASE_PATH'])
            c = conn.cursor()
            
            # 카테고리 매핑 생성 (이름 -> ID)
            c.execute('SELECT id, name FROM categories')
            category_map = {name: id for id, name in c.fetchall()}
            
            updated_count = 0
            error_rows = []
            
            # 데이터 행 처리 (첫 번째 행은 헤더이므로 2번째부터)
            for row_num in range(2, ws.max_row + 1):
                try:
                    # 셀 값 읽기
                    question_id = ws.cell(row=row_num, column=1).value
                    category_name = ws.cell(row=row_num, column=2).value
                    code = ws.cell(row=row_num, column=3).value
                    title = ws.cell(row=row_num, column=4).value
                    description = ws.cell(row=row_num, column=5).value
                    
                    # 필수 필드 검증
                    if not all([question_id, category_name, code, title]):
                        error_rows.append(f"행 {row_num}: 필수 필드 누락")
                        continue
                    
                    # 카테고리 존재 확인
                    if category_name not in category_map:
                        error_rows.append(f"행 {row_num}: 존재하지 않는 카테고리 '{category_name}'")
                        continue
                    
                    category_id = category_map[category_name]
                    
                    # 문항 존재 여부 확인
                    c.execute('SELECT id FROM questions WHERE id = ?', (question_id,))
                    existing_question = c.fetchone()
                    
                    if existing_question:
                        # 기존 문항 업데이트
                        c.execute('''UPDATE questions SET category_id = ?, code = ?, title = ?, description = ?
                                     WHERE id = ?''', (category_id, code, title, description or '', question_id))
                        action = "업데이트"
                    else:
                        # 새 문항 생성
                        # order_num은 임시로 999로 설정 (나중에 자동 정렬에서 수정됨)
                        c.execute('''INSERT INTO questions (id, category_id, code, title, description, max_score, order_num)
                                     VALUES (?, ?, ?, ?, ?, 5, 999)''', 
                                  (question_id, category_id, code, title, description or ''))
                        action = "생성"
                    
                    # 선택지 처리 (DELETE 후 INSERT 방식으로 안전하게 처리)
                    c.execute('DELETE FROM question_options WHERE question_id = ?', (question_id,))
                    
                    for score in range(1, 6):
                        option_desc = ws.cell(row=row_num, column=5 + score).value
                        if not option_desc:
                            option_desc = f"Level {score} - {['기본', '관리', '정의', '최적화', '혁신'][score-1]} 수준"
                        
                        c.execute('''INSERT INTO question_options (question_id, score, description)
                                     VALUES (?, ?, ?)''', (question_id, score, option_desc))
                    
                    updated_count += 1
                    
                except Exception as e:
                    error_rows.append(f"행 {row_num}: 처리 오류 - {str(e)}")
            
            # 문항 순서 자동 재정렬
            if updated_count > 0:
                try:
                    # 카테고리별로 문항 코드 순으로 order_num 재설정
                    c.execute('''
                        SELECT q.id, q.code, q.category_id, c.order_num as cat_order
                        FROM questions q
                        JOIN categories c ON q.category_id = c.id
                        ORDER BY c.order_num, q.code
                    ''')
                    
                    questions = c.fetchall()
                    current_category = None
                    category_order = 0
                    
                    for q_id, q_code, category_id, cat_order in questions:
                        if category_id != current_category:
                            current_category = category_id
                            category_order = 1
                        
                        c.execute('UPDATE questions SET order_num = ? WHERE id = ?', 
                                (category_order, q_id))
                        category_order += 1
                    
                    print(f"문항 순서 자동 재정렬 완료")
                except Exception as e:
                    print(f"문항 순서 재정렬 중 오류: {e}")
            
            conn.commit()
            conn.close()
            
            # 임시 파일 삭제
            os.unlink(tmp.name)
            
            # 결과 메시지
            if updated_count > 0:
                flash(f'{updated_count}개의 문항이 성공적으로 업데이트되었습니다. (순서 자동 정렬 완료)')
            
            if error_rows:
                error_msg = "다음 행에서 오류가 발생했습니다:\n" + "\n".join(error_rows[:10])
                if len(error_rows) > 10:
                    error_msg += f"\n... 외 {len(error_rows) - 10}개 오류"
                flash(error_msg)
            
    except Exception as e:
        flash(f'파일 처리 중 오류가 발생했습니다: {str(e)}')
    
    return redirect(url_for('questions'))

@app.route('/assessment/<int:assessment_id>/report')
def generate_pdf_report(assessment_id):
    """평가 결과를 PDF 보고서로 생성"""
    try:
        # 평가 데이터 조회
        conn = sqlite3.connect(app.config['DATABASE_PATH'])
        c = conn.cursor()
        
        # 기본 평가 정보
        c.execute('''SELECT a.*, co.name as company_name, co.industry, co.size
                     FROM assessments a
                     JOIN companies co ON a.company_id = co.id
                     WHERE a.id = ?''', (assessment_id,))
        assessment_data = c.fetchone()
        
        if not assessment_data:
            flash('평가 데이터를 찾을 수 없습니다.')
            return redirect(url_for('assessments'))
        
        # 카테고리별 점수
        c.execute('''SELECT cat.id, cat.name, cat.weight, SUM(ar.score) as category_score,
                            COUNT(ar.score) * 5 as max_category_score
                     FROM categories cat
                     JOIN questions q ON cat.id = q.category_id
                     JOIN assessment_results ar ON q.id = ar.question_id
                     WHERE ar.assessment_id = ?
                     GROUP BY cat.id, cat.name, cat.weight
                     ORDER BY cat.order_num''', (assessment_id,))
        category_scores = c.fetchall()
        
        # 상세 결과 (주관식 답변 포함)
        c.execute('''SELECT q.code, q.title, ar.score, qo.description, cat.name as category_name, ar.comment
                     FROM assessment_results ar
                     JOIN questions q ON ar.question_id = q.id
                     JOIN question_options qo ON q.id = qo.question_id AND ar.score = qo.score
                     JOIN categories cat ON q.category_id = cat.id
                     WHERE ar.assessment_id = ?
                     ORDER BY cat.order_num, q.order_num''', (assessment_id,))
        detailed_results = c.fetchall()
        
        conn.close()
        
        # PDF 생성
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # 스타일 설정 (한글 폰트 적용)
        styles = getSampleStyleSheet()
        
        # 한글 폰트 적용한 스타일들
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontName=KOREAN_FONT,
            fontSize=18,
            textColor=colors.darkblue,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontName=KOREAN_FONT,
            fontSize=14,
            textColor=colors.darkblue,
            spaceBefore=20,
            spaceAfter=10
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontName=KOREAN_FONT,
            fontSize=10
        )
        
        # PDF 콘텐츠 구성
        story = []
        
        # 제목
        story.append(Paragraph("APS 준비도 진단 보고서", title_style))
        story.append(Spacer(1, 20))
        
        # 기본 정보 표
        basic_info = [
            ['평가 항목', '내용'],
            ['회사명', assessment_data[7]],
            ['업종', assessment_data[8]],
            ['규모', assessment_data[9]],
            ['평가일', assessment_data[4]],
            ['총점', f"{assessment_data[2]}/140점"],
            ['성숙도 레벨', f"Level {assessment_data[3]}"]
        ]
        
        basic_table = Table(basic_info, colWidths=[2*inch, 3*inch])
        basic_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), KOREAN_FONT),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), KOREAN_FONT),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(basic_table)
        story.append(Spacer(1, 30))
        
        # 성숙도 레벨 설명
        story.append(Paragraph("성숙도 레벨 평가", heading_style))
        
        maturity_levels = [
            "Level 1 (< 40%): 기초 수준 - 체계적인 계획 수립이 필요",
            "Level 2 (40-60%): 발전 수준 - 부분적 개선이 필요",
            "Level 3 (60-80%): 우수 수준 - 전반적으로 양호한 상태",
            "Level 4 (80-91%): 최적 수준 - 일부 고도화 필요",
            "Level 5 (≥ 91%): 혁신 수준 - APS 도입 최적 상태"
        ]
        
        for level in maturity_levels:
            story.append(Paragraph(f"• {level}", normal_style))
        
        story.append(Spacer(1, 20))
        
        # 카테고리별 점수 표
        story.append(Paragraph("카테고리별 상세 점수", heading_style))
        
        category_data = [['카테고리', '획득점수', '만점', '달성률', '가중치']]
        for cat in category_scores:
            achievement_rate = (cat[3] / cat[4]) * 100
            category_data.append([
                cat[1], 
                f"{cat[3]}점", 
                f"{cat[4]}점", 
                f"{achievement_rate:.1f}%", 
                f"{cat[2]*100:.0f}%"
            ])
        
        category_table = Table(category_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        category_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), KOREAN_FONT),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), KOREAN_FONT),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(category_table)
        story.append(Spacer(1, 30))
        
        # 상세 평가 결과 (카테고리별)
        story.append(Paragraph("상세 평가 결과", heading_style))
        
        current_category = ""
        for result in detailed_results:
            if current_category != result[4]:  # 새로운 카테고리
                current_category = result[4]
                story.append(Spacer(1, 15))
                cat_style = ParagraphStyle(
                    'CategoryStyle',
                    parent=styles['Heading3'],
                    fontName=KOREAN_FONT,
                    fontSize=12,
                    textColor=colors.darkred,
                    spaceBefore=10,
                    spaceAfter=5
                )
                story.append(Paragraph(f"▶ {current_category}", cat_style))
            
            # 문항별 결과
            question_text = f"{result[0]} {result[1]} (점수: {result[2]}/5)"
            answer_text = f"선택: {result[3]}"
            
            story.append(Paragraph(question_text, normal_style))
            answer_style = ParagraphStyle(
                'AnswerStyle',
                parent=normal_style,
                fontName=KOREAN_FONT,
                fontSize=9,
                textColor=colors.darkgreen,
                leftIndent=20,
                spaceAfter=5
            )
            story.append(Paragraph(answer_text, answer_style))
            
            # 주관식 답변이 있는 경우 추가
            if result[5]:  # comment 필드
                comment_style = ParagraphStyle(
                    'CommentStyle',
                    parent=normal_style,
                    fontName=KOREAN_FONT,
                    fontSize=9,
                    textColor=colors.darkblue,
                    leftIndent=20,
                    spaceAfter=8,
                    borderColor=colors.lightgrey,
                    borderWidth=1,
                    borderPadding=5
                )
                comment_text = f"※ 상세 의견: {result[5]}"
                story.append(Paragraph(comment_text, comment_style))
        
        # 권고사항
        story.append(Spacer(1, 30))
        story.append(Paragraph("개선 권고사항", heading_style))
        
        # 성숙도 레벨에 따른 권고사항
        level = assessment_data[3]
        recommendations = []
        
        if level == 1:
            recommendations = [
                "기본적인 생산계획 프로세스 정립이 필요합니다.",
                "기준정보(BOM, 라우팅) 정확도 개선이 시급합니다.",
                "ERP 시스템 활용도를 높여야 합니다.",
                "APS 도입을 위한 기초 역량 강화가 필요합니다."
            ]
        elif level == 2:
            recommendations = [
                "생산계획 수립 주기를 단축하여 민첩성을 높이세요.",
                "실시간 데이터 수집 체계를 구축하세요.",
                "시스템 간 연동을 강화하여 정보 일관성을 확보하세요.",
                "계획 담당자의 역량 개발이 필요합니다."
            ]
        elif level == 3:
            recommendations = [
                "고급 스케줄링 기법 도입을 검토하세요.",
                "예외상황 대응 프로세스를 체계화하세요.",
                "성과 측정 및 분석 체계를 고도화하세요.",
                "APS 시스템 도입을 본격 검토할 시점입니다."
            ]
        elif level == 4:
            recommendations = [
                "AI/ML 기반 수요예측 고도화를 추진하세요.",
                "실시간 최적화 알고리즘 적용을 검토하세요.",
                "공급망 전체 관점의 통합 계획을 수립하세요.",
                "APS 시스템 도입에 최적한 상태입니다."
            ]
        else:
            recommendations = [
                "현재 우수한 수준을 유지하면서 지속적 개선을 추진하세요.",
                "차세대 기술(디지털 트윈, IoT 등) 활용을 검토하세요.",
                "벤치마킹을 통한 글로벌 수준 달성을 목표로 하세요.",
                "APS 시스템의 고도화 및 확장을 추진하세요."
            ]
        
        for rec in recommendations:
            story.append(Paragraph(f"• {rec}", normal_style))
        
        # 보고서 생성 정보
        story.append(Spacer(1, 40))
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=normal_style,
            fontName=KOREAN_FONT,
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        story.append(Paragraph(f"본 보고서는 {datetime.now().strftime('%Y년 %m월 %d일')}에 생성되었습니다.", footer_style))
        story.append(Paragraph("APS 준비도 진단 시스템 v1.0", footer_style))
        
        # PDF 생성
        doc.build(story)
        buffer.seek(0)
        
        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"APS_진단보고서_{assessment_data[7]}_{timestamp}.pdf"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/pdf'
        )
        
    except Exception as e:
        flash(f'PDF 보고서 생성 중 오류가 발생했습니다: {str(e)}')
        return redirect(url_for('assessment_detail', assessment_id=assessment_id))

if __name__ == '__main__':
    print("APS 준비도 진단 시스템을 시작합니다...")
    print("데이터베이스를 초기화합니다...")
    init_db()
    insert_initial_data()
    print("시스템이 준비되었습니다!")
    print("로컬 네트워크에서 접속 가능한 주소:")
    print("- http://localhost:5000")
    print("- http://127.0.0.1:5000")
    
    # 로컬 IP 주소 확인
    import socket
    try:
        # 로컬 IP 주소 가져오기
        hostname = socket.gethostname()
        local_ip = socket.gethostbyname(hostname)
        print(f"- http://{local_ip}:5000")
    except:
        print("- 로컬 IP 주소를 확인할 수 없습니다.")
    
    app.run(debug=True, host='0.0.0.0', port=5000)